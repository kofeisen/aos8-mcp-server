#!/usr/bin/env python3
"""Build ``aos8_mcp/data/syslog_catalog.json`` from the ArubaOS Syslog Reference Guide.

Preferred source (8.10):
  data/ArubaOS-8.10.0.0-Syslog-Reference-Guide.xlsx

Fallback (bootstrap when the xlsx is not present):
  Parses the ArubaOS 6.x Syslog Messages PDF text export bundled for development.

Usage:
  python scripts/build_syslog_catalog.py
  python scripts/build_syslog_catalog.py --xlsx /path/to/ArubaOS-8.10.0.0-Syslog-Reference-Guide.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "ArubaOS-8.10.0.0-Syslog-Reference-Guide.xlsx"
OUT_PATH = ROOT / "aos8_mcp" / "data" / "syslog_catalog.json"
FALLBACK_PDF_TEXT = (
    Path(__file__).resolve().parents[2]
    / ".cursor/projects/root-aos8-mcp-server/agent-tools/e6b394de-6658-406e-b2db-21582c6e4a8c.txt"
)

_CATEGORIES = ("System", "Security", "Wireless", "Network", "User", "ARM")
_SEVERITIES = (
    "Emergency",
    "Alert",
    "Critical",
    "Error",
    "Warning",
    "Notice",
    "Information",
    "Informational",
    "Debug",
)

_RE_ERROR_LINE = re.compile(r"^(\d{6})\s+(.+)$")
_RE_PLACEHOLDER = re.compile(r"\[[^\]]+\]")

# ArubaOS Table 2 — process → logging categories (primary = first).
_PROCESS_CATEGORIES: dict[str, list[str]] = {
    "802.1x": ["security", "network", "system", "user", "wireless"],
    "aaa": ["security", "system", "user"],
    "ads": ["system"],
    "approc": ["system"],
    "authmgr": ["security", "network", "system", "user", "wireless"],
    "amutil": ["security", "system", "user"],
    "radius": ["security", "network", "system", "user", "wireless"],
    "tacacs": ["security", "system", "user"],
    "certmgr": ["security", "system"],
    "cfgm": ["system"],
    "cli": ["system"],
    "crypto": ["security", "network", "system", "user"],
    "cts": ["system"],
    "db": ["system"],
    "dbsync": ["system"],
    "dhcpd": ["network"],
    "esi": ["system", "network", "user"],
    "fpapps": ["network", "system"],
    "httpd": ["system", "security"],
    "l2tp": ["security"],
    "ldap": ["security", "network", "system", "user", "wireless"],
    "licensemgr": ["system"],
    "localdb": ["security", "network", "system", "user", "wireless"],
    "meshd": ["security", "system", "wireless"],
    "mobileip": ["security", "network", "system", "user"],
    "nanny": ["system"],
    "ntp": ["network", "system"],
    "packetfilter": ["system"],
    "phonehome": ["network", "system"],
    "pim": ["system", "network", "user"],
    "ppp": ["security", "network", "system", "user"],
    "pppoed": ["security", "network", "system", "user"],
    "pptp": ["security", "network", "system"],
    "processes": ["system"],
    "profmgr": ["system"],
    "publisher": ["system"],
    "rfd": ["system"],
    "rfm": ["system"],
    "sapd": ["system"],
    "sapm": ["system", "wireless"],
    "snmp": ["security", "system"],
    "stm": ["security", "network", "system", "user", "wireless"],
    "syslogdwrap": ["system"],
    "traffic": ["system"],
    "voip": ["security", "network", "system", "user", "wireless"],
    "vrrpd": ["system"],
    "wms": ["security", "network", "system", "wireless"],
}


def _norm_category(name: str) -> str:
    s = name.strip().replace("_", " ").title()
    if s.endswith(" Messages"):
        s = s[: -len(" Messages")]
    aliases = {
        "Informational": "Information",
        "Info": "Information",
    }
    return aliases.get(s, s)


def _norm_severity(name: str) -> str:
    s = name.strip().title()
    if s.endswith(" Messages"):
        s = s[: -len(" Messages")]
    if s == "Informational":
        return "Information"
    return s


def _message_key(message: str) -> str:
    s = _RE_PLACEHOLDER.sub("", message)
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _is_useful_message(message: str) -> bool:
    key = _message_key(message)
    return len(key) >= 12


def _parse_pdf_text(path: Path) -> list[dict[str, str]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    category = "System"
    severity = "Error"
    entries: list[dict[str, str]] = []
    seen: set[str] = set()

    for line in lines:
        stripped = line.strip()
        cat_match = re.match(r"^(System|Security|Wireless|Network|User|ARM) Messages$", stripped)
        if cat_match:
            category = _norm_category(cat_match.group(1))
            continue
        sev_match = re.match(
            r"^(Emergency|Alert|Critical|Error|Warning|Notice|Information|Debug) Messages$",
            stripped,
        )
        if sev_match:
            severity = _norm_severity(sev_match.group(1))
            continue
        m = _RE_ERROR_LINE.match(stripped)
        if not m:
            continue
        error_id, message = m.group(1), m.group(2).strip()
        if error_id in seen:
            continue
        seen.add(error_id)
        entries.append(
            {
                "error_id": error_id,
                "category": category,
                "severity": severity,
                "message": message,
            }
        )
    return entries


def _header_map(row: list) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        key = str(cell).strip().lower().replace(" ", "_")
        out[key] = i
    return out


def _pick_col(headers: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        if c in headers:
            return headers[c]
    for k, idx in headers.items():
        for c in candidates:
            if c in k:
                return idx
    return None


def _parse_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    entries: list[dict[str, str]] = []
    seen: set[str] = set()

    for sheet in wb.worksheets:
        category = _norm_category(sheet.title)
        if category not in _CATEGORIES:
            # e.g. "About", "Index" — skip
            if not any(c.lower() in sheet.title.lower() for c in _CATEGORIES):
                continue
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        headers = _header_map(list(header_row))
        id_col = _pick_col(
            headers,
            "error_id",
            "message_id",
            "id",
            "errorid",
            "messageid",
        )
        msg_col = _pick_col(
            headers,
            "message",
            "message_text",
            "syslog_message",
            "message_and_description",
        )
        sev_col = _pick_col(headers, "severity", "level", "severity_level")
        cat_col = _pick_col(headers, "category", "log_category", "message_category")

        if id_col is None and msg_col is None:
            continue

        for row in rows:
            if not row:
                continue
            error_id = ""
            if id_col is not None and row[id_col] is not None:
                raw_id = str(row[id_col]).strip()
                m = re.search(r"(\d{6})", raw_id)
                if m:
                    error_id = m.group(1)
            message = ""
            if msg_col is not None and row[msg_col] is not None:
                message = str(row[msg_col]).strip().split("\n")[0]
            if not error_id and not message:
                continue
            if error_id and error_id in seen:
                continue
            severity = "Error"
            if sev_col is not None and row[sev_col] is not None:
                severity = _norm_severity(str(row[sev_col]))
            row_category = category
            if cat_col is not None and row[cat_col] is not None:
                row_category = _norm_category(str(row[cat_col]))
            if error_id:
                seen.add(error_id)
            entries.append(
                {
                    "error_id": error_id,
                    "category": row_category,
                    "severity": severity,
                    "message": message or f"[{error_id}]",
                }
            )
    wb.close()
    return entries


def build_catalog(*, xlsx: Path | None, pdf_text: Path | None) -> dict:
    if xlsx and xlsx.is_file():
        entries = _parse_xlsx(xlsx)
        source = xlsx.name
        version = "8.10.0.0"
    elif pdf_text and pdf_text.is_file():
        entries = _parse_pdf_text(pdf_text)
        source = pdf_text.name
        version = "6.x-bootstrap"
    else:
        raise SystemExit(
            f"No input found. Place the guide at {DEFAULT_XLSX} or pass --xlsx / --pdf-text."
        )

    by_id = {e["error_id"]: e for e in entries if e.get("error_id")}
    text_index: list[tuple[str, dict[str, str]]] = []
    for e in entries:
        if _is_useful_message(e.get("message", "")):
            text_index.append((_message_key(e["message"]), e))
    text_index.sort(key=lambda t: len(t[0]), reverse=True)

    return {
        "version": version,
        "source": source,
        "entry_count": len(entries),
        "categories": list(_CATEGORIES),
        "process_categories": _PROCESS_CATEGORIES,
        "by_id": by_id,
        "text_index": [
            {"key": k, "error_id": e["error_id"], "category": e["category"], "severity": e["severity"], "message": e["message"]}
            for k, e in text_index
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--pdf-text", type=Path, default=FALLBACK_PDF_TEXT)
    parser.add_argument("-o", "--output", type=Path, default=OUT_PATH)
    args = parser.parse_args()

    xlsx = args.xlsx if args.xlsx.is_file() else None
    catalog = build_catalog(xlsx=xlsx, pdf_text=args.pdf_text)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(catalog, ensure_ascii=False, indent=0), encoding="utf-8")
    print(
        f"Wrote {catalog['entry_count']} entries ({catalog['version']} from {catalog['source']}) "
        f"→ {args.output}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
